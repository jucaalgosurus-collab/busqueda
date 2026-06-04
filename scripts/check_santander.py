import openpyxl

new_file_path = r"C:\Users\JUAN CARLOS\Downloads\COMISIONES BANCOS 2026.xlsx"
wb = openpyxl.load_workbook(new_file_path, data_only=True)
sheet = wb['TODAS']

# Search for columns that have bank "SANTANDER"
santander_cols = []
for r in [5, 17, 29, 40]: # bank headers in sections (openpyxl is 1-based, so row 5, 17, etc.)
    if r <= sheet.max_row:
        for c in range(1, sheet.max_column + 1):
            val = sheet.cell(r, c).value
            if val is not None and "SANTANDER" in str(val).upper():
                # Let's find the section age/seguro
                age = ""
                seg = ""
                if r == 5:
                    age, seg = "Hasta 95 meses", "Si"
                elif r == 17:
                    age, seg = "Hasta 95 meses", "No"
                elif r == 29:
                    age, seg = "Desde 96 meses", "Si"
                elif r == 40:
                    age, seg = "Desde 96 meses", "No"
                
                # Check months row (r-1)
                month = None
                for col_back in range(c, 0, -1):
                    m_val = sheet.cell(r - 1, col_back).value
                    if m_val is not None:
                        month = str(m_val).strip()
                        break
                        
                santander_cols.append({
                    "col": c,
                    "row_header": r,
                    "age": age,
                    "seguro": seg,
                    "month": month
                })

print("Santander columns found:")
for sc in santander_cols:
    print(sc)
    
# Check values in those columns
for sc in santander_cols:
    col = sc["col"]
    row_header = sc["row_header"]
    print(f"\nValues in column {col} (header row {row_header}):")
    # Determine the end row for this section
    # Section 1 ends at row 12
    # Section 2 ends at row 24
    # Section 3 ends at row 35
    # Section 4 ends at row 46
    end_row = 12
    if row_header == 17:
        end_row = 24
    elif row_header == 29:
        end_row = 35
    elif row_header == 40:
        end_row = 46
        
    for r in range(row_header + 1, end_row + 1):
        val = sheet.cell(r, col).value
        label = sheet.cell(r, 1).value
        print(f"  Row {r:02d} | Label: {label} | Val: {val}")
