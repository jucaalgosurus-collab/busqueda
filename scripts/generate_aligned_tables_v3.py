import pandas as pd
import openpyxl
import os

new_file_path = r"C:\Users\JUAN CARLOS\Downloads\COMISIONES BANCOS 2026.xlsx"
old_xlsx_path = r"C:\Users\JUAN CARLOS\Desktop\Helen\Porcentajes_old_backup.xlsx"

# Load old file to preserve structural details
df_old = pd.read_excel(old_xlsx_path)

# Parse new file row-by-row
wb = openpyxl.load_workbook(new_file_path, data_only=True)
sheet = wb['TODAS']

# List to store parsed records
parsed_records = []

sections = [
    {"start": 2, "end": 12, "age": "Hasta 95 meses", "seguro": "Si"},
    {"start": 14, "end": 24, "age": "Hasta 95 meses", "seguro": "No"},
    {"start": 26, "end": 35, "age": "Desde 96 meses", "seguro": "Si"},
    {"start": 37, "end": 46, "age": "Desde 96 meses", "seguro": "No"},
]

for sec in sections:
    start = sec["start"]
    end = sec["end"]
    age = sec["age"]
    seguro = sec["seguro"]
    
    months_row = [sheet.cell(start + 2, c).value for c in range(1, 47)]
    banks_row = [sheet.cell(start + 3, c).value for c in range(1, 47)]
    
    current_month = None
    cols_mapping = {}
    for col_idx in range(1, len(months_row)):
        m = months_row[col_idx]
        b = banks_row[col_idx]
        if m is not None:
            current_month = str(m).strip()
        if b is not None:
            cols_mapping[col_idx] = (current_month, str(b).strip())
            
    # Parse data rows
    for r in range(start + 4, end + 1):
        label = sheet.cell(r, 1).value
        if label is None:
            continue
        label_str = str(label).strip()
        
        # We will extract values for each bank in this row
        # First, find which banks have non-zero/non-null values in this row
        row_banks = set()
        for col_idx, (month, bank) in cols_mapping.items():
            val = sheet.cell(r, col_idx + 1).value
            if val is not None and val != 0 and val != 0.0:
                # Get base bank
                b_clean = bank.upper()
                base_bank = None
                if "SOFINCO" in b_clean:
                    base_bank = "SOFINCO"
                elif "SABADELL" in b_clean:
                    base_bank = "SABADELL"
                elif "BBVA" in b_clean:
                    base_bank = "BBVA"
                elif "COFIDIS" in b_clean:
                    base_bank = "COFIDIS"
                elif "CAIXA" in b_clean:
                    base_bank = "CAIXA"
                elif "LENDROCK" in b_clean:
                    base_bank = "LENDROCK"
                elif "SANTANDER" in b_clean:
                    base_bank = "SANTANDER"
                if base_bank:
                    row_banks.add(base_bank)
                    
        # Now for each bank that has values in this row, create a record
        for bank in row_banks:
            # Determine TIN for this bank in this row
            tin_val = None
            if label_str in ["0.0599", "0.0599"]:
                tin_val = 0.0599
            elif "-" in label_str:
                parts = label_str.split("-")
                gen_part = parts[0].strip().replace("%", "").replace(",", ".")
                gen_tin = float(gen_part) / 100.0
                
                spec_tin = None
                for part in parts[1:]:
                    part = part.strip()
                    if bank == "BBVA" and "BBVA" in part:
                        words = part.split()
                        for w in words:
                            if "%" in w or "," in w or any(c.isdigit() for c in w):
                                w_clean = w.replace("%", "").replace(",", ".").strip()
                                try:
                                    spec_tin = float(w_clean) / 100.0
                                except:
                                    pass
                if spec_tin is not None:
                    tin_val = spec_tin
                else:
                    tin_val = gen_tin
            else:
                try:
                    val_str = label_str.replace("%", "").replace(",", ".").strip()
                    tin_val = float(val_str)
                    if tin_val > 1.0:
                        tin_val = tin_val / 100.0
                except:
                    pass
                    
            if tin_val is not None:
                tin_val = round(tin_val, 4)
                
            if tin_val is None:
                continue
                
            actual_age = age
            if bank in ["BBVA", "COFIDIS"]:
                if age == "Hasta 95 meses":
                    actual_age = "Hasta 84 meses"
                elif age == "Desde 96 meses":
                    actual_age = "Desde 84 meses"
            
            # Extract monthly values for this bank in this row
            record_vals = {"24m": "-", "36m": "-", "48m": "-", "60m": "-", "72m": "-", "84m": "-", "96-120m": "-"}
            for col_idx, (month, col_bank) in cols_mapping.items():
                cb_clean = col_bank.upper()
                col_base_bank = None
                if "SOFINCO" in cb_clean:
                    col_base_bank = "SOFINCO"
                elif "SABADELL" in cb_clean:
                    col_base_bank = "SABADELL"
                elif "BBVA" in cb_clean:
                    col_base_bank = "BBVA"
                elif "COFIDIS" in cb_clean:
                    col_base_bank = "COFIDIS"
                elif "CAIXA" in cb_clean:
                    col_base_bank = "CAIXA"
                elif "LENDROCK" in cb_clean:
                    col_base_bank = "LENDROCK"
                elif "SANTANDER" in cb_clean:
                    col_base_bank = "SANTANDER"
                    
                if col_base_bank != bank:
                    continue
                    
                val = sheet.cell(r, col_idx + 1).value
                if val is None or val == 0 or val == 0.0:
                    continue
                    
                is_96_120 = False
                if "96" in cb_clean or "120" in cb_clean:
                    is_96_120 = True
                    
                if month == "24":
                    record_vals["24m"] = val
                elif month == "36":
                    record_vals["36m"] = val
                elif month == "48":
                    record_vals["48m"] = val
                elif month == "60":
                    record_vals["60m"] = val
                elif month == "72":
                    record_vals["72m"] = val
                elif month == "84 - 120":
                    if is_96_120:
                        record_vals["96-120m"] = val
                    else:
                        if "84" in col_bank:
                            record_vals["84m"] = val
                        elif col_bank in ["SABADELL", "SOFINCO", "BBVA", "COFIDIS", "LENDROCK", "SANTANDER"]:
                            has_separate_96_120 = bank in ["SOFINCO", "CAIXA", "BBVA", "LENDROCK", "COFIDIS"]
                            if has_separate_96_120:
                                record_vals["84m"] = val
                            else:
                                record_vals["84m"] = val
                                record_vals["96-120m"] = val
                                
            parsed_records.append({
                "Entidad": bank,
                "Antigüedad": actual_age,
                "Seguro": seguro,
                "TIN Aplicado": tin_val,
                **record_vals
            })

print(f"Parsed {len(parsed_records)} bank records from new sheet.")

# Let's write a function to construct the final DataFrame combining old template and new data
def build_final_table(banks_list):
    final_rows = []
    
    # Standard bank structures (age, seguro) from old template or new entries
    bank_structures = {
        "ABANCA": [("Hasta 95 meses", "Si"), ("Hasta 95 meses", "No"), ("Desde 96 meses", "Si"), ("Desde 96 meses", "No")],
        "BBVA": [("Hasta 84 meses", "Si"), ("Hasta 84 meses", "No"), ("Desde 84 meses", "Si"), ("Desde 84 meses", "No")],
        "CAIXA": [("Hasta 95 meses", "Si"), ("Hasta 95 meses", "No")],
        "COFIDIS": [("Hasta 84 meses", "Si"), ("Hasta 84 meses", "No"), ("Desde 84 meses", "Si"), ("Desde 84 meses", "No")],
        "SABADELL": [("Hasta 95 meses", "Si"), ("Hasta 95 meses", "No"), ("Desde 96 meses", "Si"), ("Desde 96 meses", "No")],
        "SOFINCO": [("Hasta 95 meses", "Si"), ("Hasta 95 meses", "No"), ("Desde 96 meses", "Si"), ("Desde 96 meses", "No")],
        "LENDROCK": [("Hasta 95 meses", "Si"), ("Hasta 95 meses", "No"), ("Desde 96 meses", "Si"), ("Desde 96 meses", "No")],
        "SANTANDER": [("Hasta 95 meses", "Si"), ("Hasta 95 meses", "No"), ("Desde 96 meses", "Si"), ("Desde 96 meses", "No")],
    }
    
    for bank in banks_list:
        if bank not in bank_structures:
            continue
            
        for age, seguro in bank_structures[bank]:
            # Get TINs from old file
            old_tins = set(round(t, 4) for t in df_old[(df_old['Entidad'] == bank) & (df_old['Antigüedad'] == age) & (df_old['Seguro'] == seguro)]['TIN Aplicado'].tolist())
            
            # Get TINs from new parsed records
            new_tins = set(round(r["TIN Aplicado"], 4) for r in parsed_records if r["Entidad"] == bank and r["Antigüedad"] == age and r["Seguro"] == seguro)
            
            # Combine all unique TINs and sort them ascending
            combined_tins = sorted(list(old_tins.union(new_tins)))
            
            for tin in combined_tins:
                # Find if we have a parsed record in the new data for this combination
                # Note: we might have multiple parsed records due to the BBVA 5.99% duplicate!
                # Let's find all matching records.
                matches = [r for r in parsed_records if r["Entidad"] == bank and r["Antigüedad"] == age and r["Seguro"] == seguro and abs(r["TIN Aplicado"] - tin) < 1e-6]
                
                if len(matches) == 0:
                    # Not found in new sheet (e.g. ABANCA or a dropped rate like COFIDIS 11.99%)
                    # Write row with all "-"
                    final_rows.append({
                        "Entidad": bank,
                        "Antigüedad": age,
                        "Seguro": seguro,
                        "TIN Aplicado": tin,
                        "24m": "-", "36m": "-", "48m": "-", "60m": "-", "72m": "-", "84m": "-", "96-120m": "-"
                    })
                else:
                    # Write each matching record as a row (handles duplicate rows like BBVA 5.99% campaign and standard)
                    for m in matches:
                        final_rows.append({
                            "Entidad": bank,
                            "Antigüedad": age,
                            "Seguro": seguro,
                            "TIN Aplicado": tin,
                            "24m": m["24m"], "36m": m["36m"], "48m": m["48m"], "60m": m["60m"], "72m": m["72m"], "84m": m["84m"], "96-120m": m["96-120m"]
                        })
                        
    df_final = pd.DataFrame(final_rows)
    
    # Custom sorting
    def sort_key(row):
        ent = row["Entidad"]
        
        age = row["Antigüedad"]
        age_score = 0
        if "Hasta" in str(age):
            age_score = 0
        elif "Desde" in str(age):
            age_score = 1
        else:
            age_score = 2
            
        seg = row["Seguro"]
        seg_score = 0
        if seg == "Si":
            seg_score = 0
        elif seg == "No":
            seg_score = 1
        else:
            seg_score = 2
            
        tin = row["TIN Aplicado"]
        return (ent, age_score, seg_score, tin)
        
    df_final["sort_key"] = df_final.apply(sort_key, axis=1)
    df_final = df_final.sort_values(by="sort_key").drop(columns=["sort_key"]).reset_index(drop=True)
    
    return df_final

original_banks = ["ABANCA", "BBVA", "CAIXA", "COFIDIS", "SABADELL", "SOFINCO"]
all_banks = ["ABANCA", "BBVA", "CAIXA", "COFIDIS", "LENDROCK", "SABADELL", "SANTANDER", "SOFINCO"]

df_final_opt_a = build_final_table(original_banks)
df_final_opt_b = build_final_table(all_banks)

print(f"Option A (Original Banks) row count (v3): {len(df_final_opt_a)}")
print(f"Option B (All Banks) row count (v3): {len(df_final_opt_b)}")

# Save to scripts dir
df_final_opt_a.to_excel("scripts/Porcentajes_OptA_v3.xlsx", index=False)
df_final_opt_b.to_excel("scripts/Porcentajes_OptB_v3.xlsx", index=False)
df_final_opt_a.to_csv("scripts/Porcentajes_OptA_v3.csv", index=False)
df_final_opt_b.to_csv("scripts/Porcentajes_OptB_v3.csv", index=False)
print("Saved Option A and B (v3) to scripts/ dir.")
