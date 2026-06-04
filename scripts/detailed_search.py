import openpyxl
import pandas as pd
import os
import glob

files_to_search = [
    r"C:\Users\JUAN CARLOS\Downloads\COMISIONES BANCOS 2026.xlsx",
    r"C:\Users\JUAN CARLOS\Desktop\Helen\Porcentajes.xlsx",
    r"C:\Users\JUAN CARLOS\Desktop\Helen\Porcentajes.csv",
    r"C:\Users\JUAN CARLOS\Desktop\Helen\Entregas.xlsx",
    r"C:\Users\JUAN CARLOS\Desktop\Helen\Entregas.csv",
    r"C:\Users\JUAN CARLOS\Desktop\Helen\Extracciones.xlsx"
]

print("Starting detailed cell-by-cell search...")

for filepath in files_to_search:
    if not os.path.exists(filepath):
        print(f"File not found: {filepath}")
        continue
    
    print(f"\nSearching in file: {filepath}")
    
    if filepath.endswith(".xlsx"):
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            print(f"  All sheets in file: {wb.sheetnames}")
            for sheetname in wb.sheetnames:
                sheet = wb[sheetname]
                print(f"  Sheet: {sheetname} (State: {sheet.sheet_state})")
                for r in range(1, sheet.max_row + 1):
                    for c in range(1, sheet.max_column + 1):
                        val = sheet.cell(r, c).value
                        if val is not None:
                            val_str = str(val).lower()
                            if "iber" in val_str or "caja" in val_str:
                                print(f"    Match in cell ({r}, {c}): '{val}'")
        except Exception as e:
            print(f"  Error reading excel: {e}")
            
    elif filepath.endswith(".csv"):
        try:
            df = pd.read_csv(filepath, sep=None, engine='python', header=None)
            for r_idx, row in df.iterrows():
                for c_idx, val in enumerate(row):
                    if pd.notna(val):
                        val_str = str(val).lower()
                        if "iber" in val_str or "caja" in val_str:
                            print(f"    Match in row {r_idx}, col {c_idx}: '{val}'")
        except Exception as e:
            print(f"  Error reading csv: {e}")
