import pandas as pd
import openpyxl
import os

new_file_path = r"C:\Users\JUAN CARLOS\Downloads\COMISIONES BANCOS 2026.xlsx"
old_xlsx_path = r"C:\Users\JUAN CARLOS\Desktop\Helen\Porcentajes.xlsx"

# Load old file to preserve structural details
df_old = pd.read_excel(old_xlsx_path)

# Parse new file row-by-row
wb = openpyxl.load_workbook(new_file_path, data_only=True)
sheet = wb['TODAS']

# List to store parsed records
parsed_records = []

sections = [
    {"start": 2, "end": 12, "age": "Hasta 95 meses", "seguro": "Si"},
    {"start": 14, "end": 24, "age": "Hasta 95 meses", "seguro": "No"},
    {"start": 26, "end": 35, "age": "Desde 96 meses", "seguro": "Si"},
    {"start": 37, "end": 46, "age": "Desde 96 meses", "seguro": "No"},
]

for sec in sections:
    start = sec["start"]
    end = sec["end"]
    age = sec["age"]
    seguro = sec["seguro"]
    
    months_row = [sheet.cell(start + 2, c).value for c in range(1, 47)]
    banks_row = [sheet.cell(start + 3, c).value for c in range(1, 47)]
    
    current_month = None
    cols_mapping = {}
    for col_idx in range(1, len(months_row)):
        m = months_row[col_idx]
        b = banks_row[col_idx]
        if m is not None:
            current_month = str(m).strip()
        if b is not None:
            cols_mapping[col_idx] = (current_month, str(b).strip())
            
    # Parse data rows
    for r in range(start + 4, end + 1):
        label = sheet.cell(r, 1).value
        if label is None:
            continue
        label_str = str(label).strip()
        
        # We will extract values for each bank in this row
        # First, find which banks have non-zero/non-null values in this row
        row_banks = set()
        for col_idx, (month, bank) in cols_mapping.items():
            val = sheet.cell(r, col_idx + 1).value
            if val is not None and val != 0 and val != 0.0:
                # Get base bank
                b_clean = bank.upper()
                base_bank = None
                if "SOFINCO" in b_clean:
                    base_bank = "SOFINCO"
                elif "SABADELL" in b_clean:
                    base_bank = "SABADELL"
                elif "BBVA" in b_clean:
                    base_bank = "BBVA"
                elif "COFIDIS" in b_clean:
                    base_bank = "COFIDIS"
                elif "CAIXA" in b_clean:
                    base_bank = "CAIXA"
                elif "LENDROCK" in b_clean:
                    base_bank = "LENDROCK"
                elif "SANTANDER" in b_clean:
                    base_bank = "SANTANDER"
                if base_bank:
                    row_banks.add(base_bank)
                    
        # Now for each bank that has values in this row, create a record
        for bank in row_banks:
            # Determine TIN for this bank in this row
            tin_val = None
            if label_str in ["0.0599", "0.0599"]:
                tin_val = 0.0599
            elif "-" in label_str:
                parts = label_str.split("-")
                gen_part = parts[0].strip().replace("%", "").replace(",", ".")
                gen_tin = float(gen_part) / 100.0
                
                spec_tin = None
                for part in parts[1:]:
                    part = part.strip()
                    if bank == "BBVA" and "BBVA" in part:
                        words = part.split()
                        for w in words:
                            if "%" in w or "," in w or any(c.isdigit() for c in w):
                                w_clean = w.replace("%", "").replace(",", ".").strip()
                                try:
                                    spec_tin = float(w_clean) / 100.0
                                except:
                                    pass
                if spec_tin is not None:
                    tin_val = spec_tin
                else:
                    tin_val = gen_tin
            else:
                try:
                    val_str = label_str.replace("%", "").replace(",", ".").strip()
                    tin_val = float(val_str)
                    if tin_val > 1.0:
                        tin_val = tin_val / 100.0
                except:
                    pass
                    
            if tin_val is None:
                continue
                
            actual_age = age
            if bank in ["BBVA", "COFIDIS"]:
                if age == "Hasta 95 meses":
                    actual_age = "Hasta 84 meses"
                elif age == "Desde 96 meses":
                    actual_age = "Desde 84 meses"
            
            # Extract monthly values for this bank in this row
            record_vals = {"24m": "-", "36m": "-", "48m": "-", "60m": "-", "72m": "-", "84m": "-", "96-120m": "-"}
            for col_idx, (month, col_bank) in cols_mapping.items():
                # Check if this column belongs to the current bank
                cb_clean = col_bank.upper()
                col_base_bank = None
                if "SOFINCO" in cb_clean:
                    col_base_bank = "SOFINCO"
                elif "SABADELL" in cb_clean:
                    col_base_bank = "SABADELL"
                elif "BBVA" in cb_clean:
                    col_base_bank = "BBVA"
                elif "COFIDIS" in cb_clean:
                    col_base_bank = "COFIDIS"
                elif "CAIXA" in cb_clean:
                    col_base_bank = "CAIXA"
                elif "LENDROCK" in cb_clean:
                    col_base_bank = "LENDROCK"
                elif "SANTANDER" in cb_clean:
                    col_base_bank = "SANTANDER"
                    
                if col_base_bank != bank:
                    continue
                    
                val = sheet.cell(r, col_idx + 1).value
                if val is None or val == 0 or val == 0.0:
                    continue
                    
                is_96_120 = False
                if "96" in cb_clean or "120" in cb_clean:
                    is_96_120 = True
                    
                if month == "24":
                    record_vals["24m"] = val
                elif month == "36":
                    record_vals["36m"] = val
                elif month == "48":
                    record_vals["48m"] = val
                elif month == "60":
                    record_vals["60m"] = val
                elif month == "72":
                    record_vals["72m"] = val
                elif month == "84 - 120":
                    if is_96_120:
                        record_vals["96-120m"] = val
                    else:
                        if "84" in col_bank:
                            record_vals["84m"] = val
                        elif col_bank in ["SABADELL", "SOFINCO", "BBVA", "COFIDIS", "LENDROCK", "SANTANDER"]:
                            has_separate_96_120 = bank in ["SOFINCO", "CAIXA", "BBVA", "LENDROCK", "COFIDIS"]
                            if has_separate_96_120:
                                record_vals["84m"] = val
                            else:
                                record_vals["84m"] = val
                                record_vals["96-120m"] = val
                                
            parsed_records.append({
                "Entidad": bank,
                "Antigüedad": actual_age,
                "Seguro": seguro,
                "TIN Aplicado": tin_val,
                **record_vals
            })

print(f"Parsed {len(parsed_records)} bank records from new sheet.")

# Let's write a function to construct the final DataFrame
def build_final_table(banks_list):
    # We want to combine the old template structure with the new parsed records.
    # For each bank in banks_list:
    # If the bank is in the new parsed records, we extract all records for it.
    # What about ABANCA (which is not in new records)?
    # We copy its rows from df_old, but set all monthly columns to "-" (since it's not active).
    # Wait, what if the bank is in the new parsed records but also had some rows in df_old?
    # We should use the new parsed records.
    # To keep the grouping and order clean:
    # We will sort the final table by:
    # 1. Entidad (alphabetical: ABANCA, BBVA, CAIXA, COFIDIS, LENDROCK, SABADELL, SANTANDER, SOFINCO)
    # 2. Antigüedad (custom: "Hasta..." before "Desde...")
    # 3. Seguro (custom: "Si" before "No")
    # 4. TIN Aplicado (ascending)
    
    final_rows = []
    
    for bank in banks_list:
        bank_records = [r for r in parsed_records if r["Entidad"] == bank]
        
        if len(bank_records) == 0:
            # If bank is not in new records (like ABANCA), copy its rows from df_old but clear monthly values
            old_bank_rows = df_old[df_old["Entidad"] == bank].copy()
            if len(old_bank_rows) > 0:
                for col in ["24m", "36m", "48m", "60m", "72m", "84m", "96-120m"]:
                    old_bank_rows[col] = "-"
                final_rows.append(old_bank_rows)
        else:
            # Use the new parsed records
            df_bank = pd.DataFrame(bank_records)
            final_rows.append(df_bank)
            
    df_final = pd.concat(final_rows, ignore_index=True)
    
    # Custom sorting
    # Define custom sort keys
    def sort_key(row):
        # Entidad
        ent = row["Entidad"]
        
        # Antigüedad
        age = row["Antigüedad"]
        age_score = 0
        if "Hasta" in str(age):
            age_score = 0
        elif "Desde" in str(age):
            age_score = 1
        else:
            age_score = 2
            
        # Seguro
        seg = row["Seguro"]
        seg_score = 0
        if seg == "Si":
            seg_score = 0
        elif seg == "No":
            seg_score = 1
        else:
            seg_score = 2
            
        # TIN Aplicado
        tin = row["TIN Aplicado"]
        
        return (ent, age_score, seg_score, tin)
        
    # Apply sorting
    df_final["sort_key"] = df_final.apply(sort_key, axis=1)
    df_final = df_final.sort_values(by="sort_key").drop(columns=["sort_key"]).reset_index(drop=True)
    
    return df_final

original_banks = ["ABANCA", "BBVA", "CAIXA", "COFIDIS", "SABADELL", "SOFINCO"]
all_banks = ["ABANCA", "BBVA", "CAIXA", "COFIDIS", "LENDROCK", "SABADELL", "SANTANDER", "SOFINCO"]

df_final_opt_a = build_final_table(original_banks)
df_final_opt_b = build_final_table(all_banks)

print(f"Option A (Original Banks) row count: {len(df_final_opt_a)}")
print(f"Option B (All Banks) row count: {len(df_final_opt_b)}")

# Let's inspect some rows in Option A for BBVA
print("\nBBVA rows in Option A (v2):")
print(df_final_opt_a[df_final_opt_a['Entidad'] == 'BBVA'].to_string())

# Save them to script dir
df_final_opt_a.to_excel("scripts/Porcentajes_OptA_v2.xlsx", index=False)
df_final_opt_b.to_excel("scripts/Porcentajes_OptB_v2.xlsx", index=False)
df_final_opt_a.to_csv("scripts/Porcentajes_OptA_v2.csv", index=False)
df_final_opt_b.to_csv("scripts/Porcentajes_OptB_v2.csv", index=False)
print("Saved Option A and B (v2) to scripts/ dir.")
