import pandas as pd
import openpyxl

old_xlsx_path = r"C:\Users\JUAN CARLOS\Desktop\Helen\Porcentajes.xlsx"
df_old = pd.read_excel(old_xlsx_path)

# Let's extract the new data in a clean dictionary
# Structure: { (bank, age_group, seguro_val, tin_val) -> [val_24m, val_36m, val_48m, val_60m, val_72m, val_84m, val_96_120m] }
new_data = {}

wb = openpyxl.load_workbook(r"C:\Users\JUAN CARLOS\Downloads\COMISIONES BANCOS 2026.xlsx", data_only=True)
sheet = wb['TODAS']

def get_age_group(section_name):
    # Map section name to "Hasta 95 meses" / "Desde 96 meses" / "Hasta 84 meses" / "Desde 84 meses"
    s = section_name.lower()
    if "hasta 95 meses" in s:
        return "Hasta 95 meses"
    elif "desde 96 meses" in s:
        return "Desde 96 meses"
    return section_name

# Let's write a parser that maps all values
# We'll parse the 4 sections of the sheet.
sections = [
    {"start": 2, "end": 12, "age": "Hasta 95 meses", "seguro": "Si"},
    {"start": 14, "end": 24, "age": "Hasta 95 meses", "seguro": "No"},
    {"start": 26, "end": 35, "age": "Desde 96 meses", "seguro": "Si"},
    {"start": 37, "end": 46, "age": "Desde 96 meses", "seguro": "No"},
]

for sec in sections:
    start = sec["start"]
    end = sec["end"]
    age = sec["age"]
    seguro = sec["seguro"]
    
    # Months are at start + 2
    # Banks are at start + 3
    months_row = [sheet.cell(start + 2, c).value for c in range(1, 47)]
    banks_row = [sheet.cell(start + 3, c).value for c in range(1, 47)]
    
    # Build column mapping: col_idx (0-based) -> (month, bank)
    current_month = None
    cols_mapping = {}
    for col_idx in range(1, len(months_row)):
        m = months_row[col_idx]
        b = banks_row[col_idx]
        if m is not None:
            current_month = str(m).strip()
        if b is not None:
            cols_mapping[col_idx] = (current_month, str(b).strip())
            
    # Iterate over data rows
    for r in range(start + 4, end + 1):
        label = sheet.cell(r, 1).value
        if label is None:
            continue
        label_str = str(label).strip()
        
        # Now for each bank in the columns, we extract its value
        # Let's find what TIN applies to each bank in this row.
        # We can look at the label:
        # e.g., "5,99% - BBVA 5,99%", "6,99% - BBVA 7,25%", "0.0599"
        # We can parse the label to find the TIN for each bank.
        # Let's inspect the label format.
        # If it's a simple float like 0.0599, it means TIN is 0.0599 (5.99%) for whatever bank has a value.
        # If it's "5,99% - BBVA 5,99%", then BBVA TIN is 5.99%, others is 5.99%.
        # If it's "6,99% - BBVA 7,25%", then BBVA TIN is 7.25% (0.0725), others is 6.99% (0.0699).
        # Let's write a parsing function for TINs.
        
        # Let's extract values
        for col_idx, (month, bank) in cols_mapping.items():
            val = sheet.cell(r, col_idx + 1).value
            if val is None:
                continue
            
            # Determine the TIN for this bank in this row
            tin_val = None
            if label_str == "0.0599" or label_str == "0.0599":
                tin_val = 0.0599
            elif "-" in label_str:
                parts = label_str.split("-")
                # e.g., "6,99% " and " BBVA 7,25%"
                gen_part = parts[0].strip().replace("%", "").replace(",", ".")
                gen_tin = float(gen_part) / 100.0
                
                # Check if there is a bank-specific TIN
                spec_tin = None
                for part in parts[1:]:
                    part = part.strip()
                    # e.g., "BBVA 7,25%" or "BBVA/96 - 120" (wait, row 12 of section 3 has "BBVA 11,25%")
                    if bank.startswith("BBVA") and "BBVA" in part:
                        # Extract float
                        words = part.split()
                        for w in words:
                            if "%" in w or "," in w or any(c.isdigit() for c in w):
                                w_clean = w.replace("%", "").replace(",", ".").strip()
                                try:
                                    spec_tin = float(w_clean) / 100.0
                                except:
                                    pass
                if spec_tin is not None:
                    tin_val = spec_tin
                else:
                    tin_val = gen_tin
            else:
                # Try to parse label as a percentage or float
                try:
                    val_str = label_str.replace("%", "").replace(",", ".").strip()
                    tin_val = float(val_str)
                    if tin_val > 1.0:
                        tin_val = tin_val / 100.0
                except:
                    pass
            
            if tin_val is None:
                continue
            
            # Clean month name
            # "24", "36", "48", "60", "72", "84 - 120"
            m_key = month
            
            # Map bank name
            # Note that in Month 84-120 we have: "SOFINCO 96-120", "CAIXA 96 - 120 ", "BBVA/96 - 120", "LENDROCK 96-120", "COFIDIS 96 - 120"
            # And "CAIXA 84"
            # Let's clean the bank name to get the base Entidad
            b_clean = bank.upper()
            is_96_120 = False
            if "96" in b_clean or "120" in b_clean:
                is_96_120 = True
            
            base_bank = None
            if "SOFINCO" in b_clean:
                base_bank = "SOFINCO"
            elif "SABADELL" in b_clean:
                base_bank = "SABADELL"
            elif "BBVA" in b_clean:
                base_bank = "BBVA"
            elif "COFIDIS" in b_clean:
                base_bank = "COFIDIS"
            elif "CAIXA" in b_clean:
                base_bank = "CAIXA"
            elif "LENDROCK" in b_clean:
                base_bank = "LENDROCK"
            elif "SANTANDER" in b_clean:
                base_bank = "SANTANDER"
            
            if base_bank is None:
                continue
            
            # Map age group for BBVA and COFIDIS: they are "Hasta 84 meses" / "Desde 84 meses" in the old file,
            # but in the new file they are under the headers "Coches de hasta 95 meses (BBVA y COFIDIS sólo hasta 84 meses)"
            # and "Coches desde 96 meses (BBVA y COFIDIS desde 84 meses)".
            # So:
            # - For BBVA and COFIDIS, "Hasta 95 meses" maps to "Hasta 84 meses"
            # - For BBVA and COFIDIS, "Desde 96 meses" maps to "Desde 84 meses"
            # - For others, they remain "Hasta 95 meses" / "Desde 96 meses"
            actual_age = age
            if base_bank in ["BBVA", "COFIDIS"]:
                if age == "Hasta 95 meses":
                    actual_age = "Hasta 84 meses"
                elif age == "Desde 96 meses":
                    actual_age = "Desde 84 meses"
            
            key = (base_bank, actual_age, seguro, tin_val)
            if key not in new_data:
                new_data[key] = {}
            
            # Map month to column in final sheet: "24m", "36m", "48m", "60m", "72m", "84m", "96-120m"
            if m_key == "24":
                new_data[key]["24m"] = val
            elif m_key == "36":
                new_data[key]["36m"] = val
            elif m_key == "48":
                new_data[key]["48m"] = val
            elif m_key == "60":
                new_data[key]["60m"] = val
            elif m_key == "72":
                new_data[key]["72m"] = val
            elif m_key == "84 - 120":
                if is_96_120:
                    new_data[key]["96-120m"] = val
                else:
                    # If it's a specific "84" column (like CAIXA 84), it goes to 84m
                    # If it's not a 96-120 column, does it go to both 84m and 96-120m?
                    # Let's check: for SABADELL it says just "SABADELL" under "84 - 120".
                    # Let's see if SABADELL goes to both 84m and 96-120m.
                    if "84" in bank: # e.g. CAIXA 84
                        new_data[key]["84m"] = val
                    elif bank == "SABADELL" or bank == "SOFINCO" or bank == "BBVA" or bank == "COFIDIS" or bank == "LENDROCK" or bank == "SANTANDER":
                        # If there is NO separate 96-120 column or if it's the general column, let's see.
                        # Wait, SOFINCO has both "SOFINCO" and "SOFINCO 96-120".
                        # BBVA has "BBVA" and "BBVA/96 - 120".
                        # COFIDIS has "COFIDIS" and "COFIDIS 96 - 120".
                        # LENDROCK has "LENDROCK" and "LENDROCK 96-120".
                        # CAIXA has "CAIXA 84" and "CAIXA 96 - 120".
                        # SABADELL has ONLY "SABADELL"!
                        # So if the bank has a separate 96-120 column, the base column maps to 84m.
                        # If the bank does NOT have a separate 96-120 column (like SABADELL or SANTANDER),
                        # then the value applies to both 84m and 96-120m!
                        has_separate_96_120 = base_bank in ["SOFINCO", "CAIXA", "BBVA", "LENDROCK", "COFIDIS"]
                        if has_separate_96_120:
                            new_data[key]["84m"] = val
                        else:
                            new_data[key]["84m"] = val
                            new_data[key]["96-120m"] = val

print(f"Parsed {len(new_data)} unique keys from new data.")

# Let's check some examples of new data
for k, v in list(new_data.items())[:15]:
    print(f"Key: {k} -> {v}")
