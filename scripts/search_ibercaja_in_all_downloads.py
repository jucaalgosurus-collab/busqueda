import os
import glob
import pandas as pd
import openpyxl
import zipfile
import re

downloads_dir = r"C:\Users\JUAN CARLOS\Downloads"
print(f"Scanning all files in {downloads_dir} for 'ibercaja'...")

# We will try to read different file types
for filename in os.listdir(downloads_dir):
    filepath = os.path.join(downloads_dir, filename)
    if not os.path.isfile(filepath):
        continue
    
    # Skip temporary files
    if filename.startswith("~$"):
        continue
        
    ext = os.path.splitext(filename)[1].lower()
    
    # Search in text files, csv, html, sql, conf, py
    if ext in [".txt", ".csv", ".html", ".sql", ".conf", ".py", ".json", ".md"]:
        try:
            with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
                if "ibercaja" in content.lower():
                    print(f"Found in text file: {filename}")
        except Exception as e:
            pass
            
    # Search in XLSX, XLSM
    elif ext in [".xlsx", ".xlsm"]:
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            for sheetname in wb.sheetnames:
                sheet = wb[sheetname]
                for r in range(1, sheet.max_row + 1):
                    for c in range(1, sheet.max_column + 1):
                        val = sheet.cell(r, c).value
                        if val is not None and "ibercaja" in str(val).lower():
                            print(f"Found in excel {filename}, sheet {sheetname}, cell ({r},{c}): '{val}'")
        except Exception as e:
            pass
            
    # Search in XLS
    elif ext == ".xls":
        try:
            xl = pd.ExcelFile(filepath)
            for sheetname in xl.sheet_names:
                df = xl.parse(sheetname, header=None)
                for col in df.columns:
                    for idx, val in df[col].items():
                        if pd.notna(val) and "ibercaja" in str(val).lower():
                            print(f"Found in xls {filename}, sheet {sheetname}, row {idx}, col {col}: '{val}'")
        except Exception as e:
            pass
            
    # Search in ZIP
    elif ext == ".zip":
        try:
            with zipfile.ZipFile(filepath, 'r') as z:
                for name in z.namelist():
                    if "ibercaja" in name.lower():
                        print(f"Found in zip filename {filename}: {name}")
                    if name.endswith((".xlsx", ".xlsm", ".xls", ".csv", ".txt", ".xml")):
                        try:
                            with z.open(name) as f:
                                # if it's xml or txt, read it
                                if name.endswith((".txt", ".xml", ".csv")):
                                    content = f.read().decode('utf-8', errors='ignore')
                                    if "ibercaja" in content.lower():
                                        print(f"Found in zip file {filename} -> text {name}")
                        except:
                            pass
        except Exception as e:
            pass
            
    # Search in PDF
    elif ext == ".pdf":
        try:
            # We can install pypdf or pdfplumber if needed, but let's check if we can read raw text of pdf as a quick check
            with open(filepath, 'rb') as f:
                content = f.read().lower()
                if b"ibercaja" in content:
                    print(f"Found (raw bytes match) in PDF: {filename}")
        except:
            pass
            
    # Search in DOCX
    elif ext == ".docx":
        try:
            # zip search inside docx
            with zipfile.ZipFile(filepath, 'r') as z:
                # docx is a zip file containing xml files. Search for 'ibercaja' in word/document.xml
                for name in z.namelist():
                    if "document.xml" in name:
                        with z.open(name) as f:
                            content = f.read().decode('utf-8', errors='ignore').lower()
                            if "ibercaja" in content:
                                print(f"Found in DOCX {filename}")
        except:
            pass

print("Scan completed.")
