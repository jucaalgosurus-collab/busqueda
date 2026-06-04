import openpyxl

new_file_path = r"C:\Users\JUAN CARLOS\Downloads\COMISIONES BANCOS 2026.xlsx"
wb = openpyxl.load_workbook(new_file_path, data_only=True)
sheet = wb['TODAS']

# Let's run a simplified version of the parser just for CAIXA
# and print what is extracted at row 18 (SIN SEGURO, index 18)

sections = [
    {"start": 2, "end": 12, "age": "Hasta 95 meses", "seguro": "Si"},
    {"start": 14, "end": 24, "age": "Hasta 95 meses", "seguro": "No"},
]

for sec in sections:
    start = sec["start"]
    end = sec["end"]
    age = sec["age"]
    seguro = sec["seguro"]
    
    months_row = [sheet.cell(start + 2, c).value for c in range(1, 47)]
    banks_row = [sheet.cell(start + 3, c).value for c in range(1, 47)]
    
    current_month = None
    cols_mapping = {}
    for col_idx in range(1, len(months_row)):
        m = months_row[col_idx]
        b = banks_row[col_idx]
        if m is not None:
            current_month = str(m).strip()
        if b is not None:
            cols_mapping[col_idx] = (current_month, str(b).strip())
            
    for r in range(start + 4, end + 1):
        label = sheet.cell(r, 1).value
        if label is None:
            continue
        label_str = str(label).strip()
        
        # Check CAIXA values in this row
        caixa_vals = {}
        for col_idx, (month, bank) in cols_mapping.items():
            if "CAIXA" in bank.upper():
                val = sheet.cell(r, col_idx + 1).value
                if val is not None:
                    caixa_vals[(month, bank)] = val
        if len(caixa_vals) > 0:
            print(f"Row {r:02d} ({seguro}) label: '{label_str}' | CAIXA values: {caixa_vals}")
