import pandas as pd
import openpyxl
import os

new_file_path = r"C:\Users\JUAN CARLOS\Downloads\COMISIONES BANCOS 2026.xlsx"
old_xlsx_path = r"C:\Users\JUAN CARLOS\Desktop\Helen\Porcentajes.xlsx"

# 1. Parse new file
wb = openpyxl.load_workbook(new_file_path, data_only=True)
sheet = wb['TODAS']

# We will store parsed data in a dict:
# key: (bank_name, age_group, seguro_val, tin_val) -> { '24m': val, ... }
new_data = {}

sections = [
    {"start": 2, "end": 12, "age": "Hasta 95 meses", "seguro": "Si"},
    {"start": 14, "end": 24, "age": "Hasta 95 meses", "seguro": "No"},
    {"start": 26, "end": 35, "age": "Desde 96 meses", "seguro": "Si"},
    {"start": 37, "end": 46, "age": "Desde 96 meses", "seguro": "No"},
]

for sec in sections:
    start = sec["start"]
    end = sec["end"]
    age = sec["age"]
    seguro = sec["seguro"]
    
    months_row = [sheet.cell(start + 2, c).value for c in range(1, 47)]
    banks_row = [sheet.cell(start + 3, c).value for c in range(1, 47)]
    
    current_month = None
    cols_mapping = {}
    for col_idx in range(1, len(months_row)):
        m = months_row[col_idx]
        b = banks_row[col_idx]
        if m is not None:
            current_month = str(m).strip()
        if b is not None:
            cols_mapping[col_idx] = (current_month, str(b).strip())
            
    for r in range(start + 4, end + 1):
        label = sheet.cell(r, 1).value
        if label is None:
            continue
        label_str = str(label).strip()
        
        for col_idx, (month, bank) in cols_mapping.items():
            val = sheet.cell(r, col_idx + 1).value
            if val is None:
                continue
            
            # Determine TIN for this bank in this row
            tin_val = None
            if label_str in ["0.0599", "0.0599"]:
                tin_val = 0.0599
            elif "-" in label_str:
                parts = label_str.split("-")
                gen_part = parts[0].strip().replace("%", "").replace(",", ".")
                gen_tin = float(gen_part) / 100.0
                
                spec_tin = None
                for part in parts[1:]:
                    part = part.strip()
                    if bank.startswith("BBVA") and "BBVA" in part:
                        words = part.split()
                        for w in words:
                            if "%" in w or "," in w or any(c.isdigit() for c in w):
                                w_clean = w.replace("%", "").replace(",", ".").strip()
                                try:
                                    spec_tin = float(w_clean) / 100.0
                                except:
                                    pass
                if spec_tin is not None:
                    tin_val = spec_tin
                else:
                    tin_val = gen_tin
            else:
                try:
                    val_str = label_str.replace("%", "").replace(",", ".").strip()
                    tin_val = float(val_str)
                    if tin_val > 1.0:
                        tin_val = tin_val / 100.0
                except:
                    pass
            
            if tin_val is None:
                continue
            
            b_clean = bank.upper()
            is_96_120 = False
            if "96" in b_clean or "120" in b_clean:
                is_96_120 = True
            
            base_bank = None
            if "SOFINCO" in b_clean:
                base_bank = "SOFINCO"
            elif "SABADELL" in b_clean:
                base_bank = "SABADELL"
            elif "BBVA" in b_clean:
                base_bank = "BBVA"
            elif "COFIDIS" in b_clean:
                base_bank = "COFIDIS"
            elif "CAIXA" in b_clean:
                base_bank = "CAIXA"
            elif "LENDROCK" in b_clean:
                base_bank = "LENDROCK"
            elif "SANTANDER" in b_clean:
                base_bank = "SANTANDER"
            
            if base_bank is None:
                continue
            
            actual_age = age
            if base_bank in ["BBVA", "COFIDIS"]:
                if age == "Hasta 95 meses":
                    actual_age = "Hasta 84 meses"
                elif age == "Desde 96 meses":
                    actual_age = "Desde 84 meses"
            
            key = (base_bank, actual_age, seguro, tin_val)
            if key not in new_data:
                new_data[key] = {}
            
            if month == "24":
                new_data[key]["24m"] = val
            elif month == "36":
                new_data[key]["36m"] = val
            elif month == "48":
                new_data[key]["48m"] = val
            elif month == "60":
                new_data[key]["60m"] = val
            elif month == "72":
                new_data[key]["72m"] = val
            elif month == "84 - 120":
                if is_96_120:
                    new_data[key]["96-120m"] = val
                else:
                    if "84" in bank:
                        new_data[key]["84m"] = val
                    elif bank in ["SABADELL", "SOFINCO", "BBVA", "COFIDIS", "LENDROCK", "SANTANDER"]:
                        has_separate_96_120 = base_bank in ["SOFINCO", "CAIXA", "BBVA", "LENDROCK", "COFIDIS"]
                        if has_separate_96_120:
                            new_data[key]["84m"] = val
                        else:
                            new_data[key]["84m"] = val
                            new_data[key]["96-120m"] = val

# Let's inspect unique keys in new data
print(f"Parsed {len(new_data)} unique keys from new data.")

# 2. Load old file to see what rows are expected
df_old = pd.read_excel(old_xlsx_path)
print(f"Old file rows: {len(df_old)}")

# We will build rows for Option A and Option B.
# For Option A (original 6 banks: ABANCA, BBVA, CAIXA, COFIDIS, SABADELL, SOFINCO)
# We want to build the table containing all rows for these banks.
# For each bank, age, insurance, what TINs should we include?
# We should include all TINs that were in the old file, plus any new TINs that are in the new file.
# Let's list the banks for Option A:
original_banks = ["ABANCA", "BBVA", "CAIXA", "COFIDIS", "SABADELL", "SOFINCO"]
all_banks = ["ABANCA", "BBVA", "CAIXA", "COFIDIS", "LENDROCK", "SABADELL", "SANTANDER", "SOFINCO"]

# Let's write a function to construct rows for a given set of banks
def build_rows_for_banks(banks_list):
    rows = []
    
    # We want to determine what ages and insurances each bank has.
    # In the old file:
    # ABANCA: Hasta 95 meses (Si/No), Desde 96 meses (Si/No)
    # BBVA: Hasta 84 meses (Si/No), Desde 84 meses (Si/No)
    # CAIXA: Hasta 95 meses (Si/No)
    # COFIDIS: Hasta 84 meses (Si/No), Desde 84 meses (Si/No)
    # SABADELL: Hasta 95 meses (Si/No), Desde 96 meses (Si/No)
    # SOFINCO: Hasta 95 meses (Si/No), Desde 96 meses (Si/No)
    
    # For Lendrock and Santander, we'll use:
    # LENDROCK: Hasta 95 meses (Si/No), Desde 96 meses (Si/No)
    # SANTANDER: Hasta 95 meses (Si/No), Desde 96 meses (Si/No)
    
    # Let's collect all possible keys (Bank, Age, Seguro)
    bank_structures = {
        "ABANCA": [("Hasta 95 meses", "Si"), ("Hasta 95 meses", "No"), ("Desde 96 meses", "Si"), ("Desde 96 meses", "No")],
        "BBVA": [("Hasta 84 meses", "Si"), ("Hasta 84 meses", "No"), ("Desde 84 meses", "Si"), ("Desde 84 meses", "No")],
        "CAIXA": [("Hasta 95 meses", "Si"), ("Hasta 95 meses", "No")],
        "COFIDIS": [("Hasta 84 meses", "Si"), ("Hasta 84 meses", "No"), ("Desde 84 meses", "Si"), ("Desde 84 meses", "No")],
        "SABADELL": [("Hasta 95 meses", "Si"), ("Hasta 95 meses", "No"), ("Desde 96 meses", "Si"), ("Desde 96 meses", "No")],
        "SOFINCO": [("Hasta 95 meses", "Si"), ("Hasta 95 meses", "No"), ("Desde 96 meses", "Si"), ("Desde 96 meses", "No")],
        "LENDROCK": [("Hasta 95 meses", "Si"), ("Hasta 95 meses", "No"), ("Desde 96 meses", "Si"), ("Desde 96 meses", "No")],
        "SANTANDER": [("Hasta 95 meses", "Si"), ("Hasta 95 meses", "No"), ("Desde 96 meses", "Si"), ("Desde 96 meses", "No")],
    }
    
    for bank in banks_list:
        if bank not in bank_structures:
            continue
        
        for age, seguro in bank_structures[bank]:
            # What TINs should we include for this combination?
            # We should collect all TINs from the old file for this combination,
            # and all TINs from the new file for this combination.
            
            old_tins = set(df_old[(df_old['Entidad'] == bank) & (df_old['Antigüedad'] == age) & (df_old['Seguro'] == seguro)]['TIN Aplicado'].tolist())
            
            new_tins = set(k[3] for k in new_data.keys() if k[0] == bank and k[1] == age and k[2] == seguro)
            
            # Combine and sort TINs
            combined_tins = sorted(list(old_tins.union(new_tins)))
            
            for tin in combined_tins:
                # Find values
                # First check if we have new data
                new_vals = new_data.get((bank, age, seguro, tin))
                
                # Create row
                row_dict = {
                    "Entidad": bank,
                    "Antigüedad": age,
                    "Seguro": seguro,
                    "TIN Aplicado": tin
                }
                
                months = ["24m", "36m", "48m", "60m", "72m", "84m", "96-120m"]
                for m in months:
                    val = None
                    if new_vals is not None:
                        val = new_vals.get(m)
                    
                    if val is None or val == 0 or val == 0.0:
                        # Wait, what if the value was non-zero in the old file and we don't have new data for this TIN?
                        # If bank is ABANCA, we don't have new data, so we can set to '-' or keep old?
                        # The user says "Los datos son correctos pero para poder pasarla por el programa de reconocimiento necesito esten ordenadas como la tabla vieja"
                        # This implies we want the *new* 2026 data.
                        # If a rate is not in the new sheet, it means it is not offered anymore (commission is -).
                        # Let's set it to '-' (which is what the old file does).
                        row_dict[m] = "-"
                    else:
                        row_dict[m] = val
                
                rows.append(row_dict)
                
    return pd.DataFrame(rows)

df_opt_a = build_rows_for_banks(original_banks)
df_opt_b = build_rows_for_banks(all_banks)

print(f"\nOption A (Original Banks only) row count: {len(df_opt_a)}")
print(f"Option B (All Banks) row count: {len(df_opt_b)}")

# Let's check some rows in option A and option B for BBVA
print("\nBBVA rows in Option A:")
print(df_opt_a[df_opt_a['Entidad'] == 'BBVA'].to_string())

# Save them to script dir first
df_opt_a.to_excel("scripts/Porcentajes_OptA.xlsx", index=False)
df_opt_b.to_excel("scripts/Porcentajes_OptB.xlsx", index=False)
df_opt_a.to_csv("scripts/Porcentajes_OptA.csv", index=False)
df_opt_b.to_csv("scripts/Porcentajes_OptB.csv", index=False)
print("Saved Option A and B to scripts/ dir.")
